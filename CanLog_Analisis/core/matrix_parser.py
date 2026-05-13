"""
矩阵解析模块 (Matrix Parser)
负责解析 .xlsx 格式的 CAN/LIN 通信矩阵文件。

支持两种格式:
  1. canmatrix 标准格式 (DBC/ARXML/KCD 嵌入 xlsx)
  2. 广汽自定义通信矩阵模板 (GAC CMX 格式)
     - 自动检测: 如果 canmatrix 解析失败或结果为空，
       则尝试按 GAC CMX 模板解析
"""
import os
import re
import traceback
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

from logger.log_manager import logger


# ─────────────────────────────────────────────────────────────
# 数据结构（不变）
# ─────────────────────────────────────────────────────────────

@dataclass
class SignalInfo:
    """信号描述"""
    name: str
    start_bit: int
    bit_length: int
    is_signed: bool = False
    factor: float = 1.0
    offset: float = 0.0
    min_val: float = 0.0
    max_val: float = 0.0
    unit: str = ""
    comment: str = ""
    # E2E 相关标记
    is_e2e_counter: bool = False
    is_e2e_crc: bool = False


@dataclass
class MessageInfo:
    """报文描述"""
    name: str
    arbitration_id: int
    dlc: int
    cycle_time_ms: Optional[float] = None
    channel: str = ""
    sender: str = ""
    is_extended_id: bool = False
    is_fd: bool = False
    signals: List[SignalInfo] = field(default_factory=list)
    # E2E 保护信息
    e2e_enabled: bool = False
    e2e_profile: str = ""
    e2e_counter_signal: Optional[str] = None
    e2e_crc_signal: Optional[str] = None
    comment: str = ""


@dataclass
class MatrixData:
    """解析后的矩阵数据汇总"""
    file_path: str = ""
    channels: List[str] = field(default_factory=list)
    messages: Dict[int, MessageInfo] = field(default_factory=dict)
    messages_by_channel: Dict[str, List[MessageInfo]] = field(default_factory=dict)
    total_messages: int = 0
    total_signals: int = 0


# ─────────────────────────────────────────────────────────────
# 主解析器
# ─────────────────────────────────────────────────────────────

class MatrixParser:
    """
    CAN 通信矩阵解析器。
    支持 canmatrix 标准格式和广汽 GAC CMX 自定义格式。

    解析策略:
      1. 尝试 canmatrix.formats.loadp()
      2. 如果失败或解析结果为空，回退到 GAC CMX 模板解析
    """

    def __init__(self):
        self._matrix_data: Optional[MatrixData] = None

    @property
    def data(self) -> Optional[MatrixData]:
        return self._matrix_data

    def parse(self, file_path: str) -> MatrixData:
        """
        解析矩阵文件并返回结构化数据。

        Args:
            file_path: .xlsx 矩阵文件路径

        Returns:
            MatrixData 解析结果

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式错误或解析失败
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"矩阵文件不存在: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的矩阵文件格式: {ext}，请使用 .xlsx 格式")

        logger.info(f"[MatrixParser] 开始解析矩阵文件: {file_path}")

        # ── 策略 1: 尝试 canmatrix 标准解析 ──
        matrix_data = self._try_canmatrix(file_path)

        if matrix_data and matrix_data.total_messages > 0:
            logger.info(
                f"[MatrixParser] canmatrix 解析成功: "
                f"{matrix_data.total_messages} 条报文"
            )
            self._matrix_data = matrix_data
            return matrix_data

        # ── 策略 2: 回退到广汽 GAC CMX 自定义格式 ──
        logger.info(
            "[MatrixParser] canmatrix 解析失败或无数据，"
            "尝试按 GAC CMX 自定义格式解析..."
        )
        matrix_data = self._parse_gac_cmx(file_path)
        self._matrix_data = matrix_data
        return matrix_data

    # ═══════════════════════════════════════════════════════
    #  策略 1: canmatrix 标准解析
    # ═══════════════════════════════════════════════════════

    @staticmethod
    def _try_canmatrix(file_path: str) -> Optional[MatrixData]:
        """尝试用 canmatrix 解析，失败返回 None"""
        try:
            import canmatrix
            import canmatrix.formats

            db_dict = canmatrix.formats.loadp(file_path)
        except Exception as e:
            logger.warning(
                f"[MatrixParser] canmatrix 加载失败: {e}"
            )
            return None

        matrix_data = MatrixData(file_path=file_path)
        total_signals = 0

        for channel_name, db in db_dict.items():
            if not db.frames:
                continue

            ch_name = str(channel_name) if channel_name else "CAN"
            if ch_name not in matrix_data.channels:
                matrix_data.channels.append(ch_name)

            matrix_data.messages_by_channel.setdefault(ch_name, [])

            for frame in db.frames:
                msg_info = MatrixParser._parse_canmatrix_frame(frame, ch_name)
                total_signals += len(msg_info.signals)
                matrix_data.messages[msg_info.arbitration_id] = msg_info
                matrix_data.messages_by_channel[ch_name].append(msg_info)

        matrix_data.total_messages = len(matrix_data.messages)
        matrix_data.total_signals = total_signals
        return matrix_data

    @staticmethod
    def _parse_canmatrix_frame(frame, channel: str) -> MessageInfo:
        """解析 canmatrix Frame 对象"""
        cycle_time = None
        for attr_name in ("GenMsgCycleTime", "CycleTime"):
            try:
                ct_attr = frame.attribute(attr_name)
                if ct_attr is not None:
                    cycle_time = float(ct_attr)
                    break
            except (KeyError, ValueError, TypeError):
                pass

        sender = ""
        if frame.transmitters:
            sender = frame.transmitters[0]

        msg_info = MessageInfo(
            name=frame.name,
            arbitration_id=frame.arbitration_id.id,
            dlc=frame.size,
            cycle_time_ms=cycle_time,
            channel=channel,
            sender=sender,
            is_extended_id=frame.arbitration_id.extended,
            comment=frame.comment or "",
        )

        for signal in frame.signals:
            sig_info = SignalInfo(
                name=signal.name,
                start_bit=signal.start_bit,
                bit_length=signal.size,
                is_signed=signal.is_signed,
                factor=float(signal.factor) if signal.factor else 1.0,
                offset=float(signal.offset) if signal.offset else 0.0,
                min_val=float(signal.min) if signal.min else 0.0,
                max_val=float(signal.max) if signal.max else 0.0,
                unit=signal.unit or "",
                comment=signal.comment or "",
            )
            msg_info.signals.append(sig_info)
            MatrixParser._detect_e2e_signal(sig_info, msg_info)

        return msg_info

    # ═══════════════════════════════════════════════════════
    #  策略 2: 广汽 GAC CMX 自定义格式解析
    # ═══════════════════════════════════════════════════════

    # GAC CMX 列索引常量
    _COL_ECU = 0           # ECU (Tx)
    _COL_MSG_NAME = 1      # Msg Name
    _COL_MSG_ID = 2        # Msg ID (hex)
    _COL_MSG_DLC = 3       # Msg Length (bytes)
    _COL_SEND_TYPE = 4     # Msg Send Type
    _COL_CYCLE_TIME = 5    # Msg Cycle Time (ms)
    _COL_SIG_NAME = 8      # Signal Name
    _COL_SIG_COMMENT = 9   # Signal Comment
    _COL_START_BIT = 10    # Start Bit Position
    _COL_SIG_LENGTH = 11   # Signal Length
    _COL_SIG_MIN = 12      # Signal Min Value (phys)
    _COL_SIG_MAX = 13      # Signal Max Value (phys)
    _COL_RESOLUTION = 17   # Resolution (factor)
    _COL_OFFSET = 18       # Offset
    _COL_UNIT = 19         # Signal Unit

    # 需要解析的 Sheet 名前缀
    _TX_RX_PREFIXES = ("Tx_", "Rx_")

    def _parse_gac_cmx(self, file_path: str) -> MatrixData:
        """
        解析广汽 GAC CMX 格式的矩阵文件。

        结构特征:
          - Sheet 名以 "Tx_" 或 "Rx_" 开头的包含报文定义
          - Row 1: Sheet 标题
          - Row 2: 列头
          - Row 3: 空行或子标题
          - Row 4+: 数据行
          - 报文首行: Col 0 (ECU) 和 Col 1 (Msg Name) 有值
          - 信号续行: Col 0-7 为空，Col 8+ 有信号数据
        """
        logger.info(f"[MatrixParser] GAC CMX 模式: 加载 {file_path}")

        try:
            wb = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as e:
            raise ValueError(f"无法打开矩阵文件: {e}") from e

        matrix_data = MatrixData(file_path=file_path)
        total_signals = 0

        # 筛选 Tx_/Rx_ Sheet
        target_sheets = [
            sn for sn in wb.sheetnames
            if any(sn.startswith(p) for p in self._TX_RX_PREFIXES)
        ]

        logger.info(
            f"[MatrixParser] GAC CMX: 找到 {len(target_sheets)} 个"
            f" Tx/Rx Sheet: {target_sheets}"
        )

        for sheet_name in target_sheets:
            try:
                ch_name = self._extract_channel_name(sheet_name)
                msg_count, sig_count = self._parse_gac_sheet(
                    wb[sheet_name], sheet_name, ch_name, matrix_data
                )
                total_signals += sig_count
                logger.info(
                    f"[MatrixParser] Sheet '{sheet_name}': "
                    f"{msg_count} 报文, {sig_count} 信号 → Channel='{ch_name}'"
                )
            except Exception as e:
                logger.warning(
                    f"[MatrixParser] Sheet '{sheet_name}' 解析失败: {e}\n"
                    f"{traceback.format_exc()}"
                )

        wb.close()

        matrix_data.total_messages = len(matrix_data.messages)
        matrix_data.total_signals = total_signals

        if matrix_data.total_messages == 0:
            raise ValueError(
                "矩阵文件解析失败: 未找到任何有效报文。"
                "请确认文件格式是否正确。"
            )

        logger.info(
            f"[MatrixParser] GAC CMX 解析完成: "
            f"{matrix_data.total_messages} 条报文, "
            f"{matrix_data.total_signals} 个信号, "
            f"Channels: {matrix_data.channels}"
        )
        return matrix_data

    def _parse_gac_sheet(
        self,
        ws,
        sheet_name: str,
        channel: str,
        matrix_data: MatrixData,
    ) -> Tuple[int, int]:
        """
        解析单个 GAC CMX Sheet。

        Returns:
            (新增报文数, 新增信号数)
        """
        if channel not in matrix_data.channels:
            matrix_data.channels.append(channel)
        matrix_data.messages_by_channel.setdefault(channel, [])

        msg_count = 0
        sig_count = 0
        current_msg: Optional[MessageInfo] = None

        # 数据从 Row 4 开始（Row 1=标题, Row 2=列头, Row 3=空）
        # 但有些 Sheet 从 Row 3 就有数据，我们从 Row 3 开始扫描
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            vals = list(row)

            # 确保有足够的列
            while len(vals) < 20:
                vals.append(None)

            msg_name = self._safe_str(vals[self._COL_MSG_NAME])
            msg_id_str = self._safe_str(vals[self._COL_MSG_ID])
            sig_name = self._safe_str(vals[self._COL_SIG_NAME])

            # ── 检查是否是报文首行 ──
            if msg_name and msg_id_str:
                # 新报文开始
                arb_id = self._parse_msg_id(msg_id_str)
                if arb_id is None:
                    logger.debug(
                        f"[MatrixParser] Sheet '{sheet_name}' Row {row_idx}: "
                        f"无效 Msg ID '{msg_id_str}', 跳过"
                    )
                    current_msg = None
                    continue

                dlc = self._safe_int(vals[self._COL_MSG_DLC], default=8)
                cycle_time = self._safe_float(vals[self._COL_CYCLE_TIME])
                sender = self._safe_str(vals[self._COL_ECU])
                send_type = self._safe_str(vals[self._COL_SEND_TYPE])

                # 判断是否 CANFD
                is_fd = dlc > 8 or "CANFD" in sheet_name.upper()

                current_msg = MessageInfo(
                    name=msg_name,
                    arbitration_id=arb_id,
                    dlc=dlc,
                    cycle_time_ms=cycle_time,
                    channel=channel,
                    sender=sender,
                    is_fd=is_fd,
                )

                logger.debug(
                    f"[MatrixParser] 新报文: {msg_name} "
                    f"(0x{arb_id:03X}), DLC={dlc}, "
                    f"Cycle={cycle_time}ms, Sender={sender}, "
                    f"Type={send_type}"
                )

                # 如果 arb_id 已存在（Rx/Tx 都有同一报文），合并信号
                if arb_id in matrix_data.messages:
                    current_msg = matrix_data.messages[arb_id]
                    logger.debug(
                        f"[MatrixParser] 报文 0x{arb_id:03X} 已存在，"
                        f"合并信号到现有报文"
                    )
                else:
                    matrix_data.messages[arb_id] = current_msg
                    matrix_data.messages_by_channel[channel].append(current_msg)
                    msg_count += 1

                # 报文首行也可能包含第一个信号
                if sig_name:
                    sig_info = self._parse_gac_signal(vals, row_idx, sheet_name)
                    if sig_info:
                        current_msg.signals.append(sig_info)
                        self._detect_e2e_signal(sig_info, current_msg)
                        sig_count += 1

            elif sig_name and current_msg is not None:
                # ── 信号续行（属于当前报文的后续信号） ──
                sig_info = self._parse_gac_signal(vals, row_idx, sheet_name)
                if sig_info:
                    current_msg.signals.append(sig_info)
                    self._detect_e2e_signal(sig_info, current_msg)
                    sig_count += 1

        return msg_count, sig_count

    def _parse_gac_signal(
        self,
        vals: list,
        row_idx: int,
        sheet_name: str,
    ) -> Optional[SignalInfo]:
        """解析 GAC CMX 格式的单个信号行"""
        sig_name = self._safe_str(vals[self._COL_SIG_NAME])
        if not sig_name:
            return None

        start_bit = self._safe_int(vals[self._COL_START_BIT])
        bit_length = self._safe_int(vals[self._COL_SIG_LENGTH])

        if start_bit is None or bit_length is None or bit_length <= 0:
            logger.debug(
                f"[MatrixParser] Sheet '{sheet_name}' Row {row_idx}: "
                f"信号 '{sig_name}' start_bit/length 无效, 跳过"
            )
            return None

        factor = self._safe_float(vals[self._COL_RESOLUTION], default=1.0)
        offset = self._safe_float(vals[self._COL_OFFSET], default=0.0)
        min_val = self._safe_float(vals[self._COL_SIG_MIN], default=0.0)
        max_val = self._safe_float(vals[self._COL_SIG_MAX], default=0.0)
        unit = self._safe_str(vals[self._COL_UNIT])
        comment = self._safe_str(vals[self._COL_SIG_COMMENT])

        sig_info = SignalInfo(
            name=sig_name,
            start_bit=start_bit,
            bit_length=bit_length,
            factor=factor,
            offset=offset,
            min_val=min_val,
            max_val=max_val,
            unit=unit,
            comment=comment,
        )

        logger.debug(
            f"[MatrixParser]   信号: {sig_name} "
            f"(SB={start_bit}, Len={bit_length}, "
            f"Factor={factor}, Offset={offset})"
        )

        return sig_info

    # ═══════════════════════════════════════════════════════
    #  辅助方法
    # ═══════════════════════════════════════════════════════

    @staticmethod
    def _detect_e2e_signal(sig_info: SignalInfo, msg_info: MessageInfo):
        """根据信号名称自动识别 E2E Counter/CRC 信号"""
        name_lower = sig_info.name.lower()

        # 匹配 Counter 关键词
        if any(kw in name_lower for kw in (
            "counter", "alivecnt", "alive_cnt", "rollingcnt",
            "rolling_cnt", "rollcnt", "e2e_cnt",
        )):
            sig_info.is_e2e_counter = True
            msg_info.e2e_counter_signal = sig_info.name
            msg_info.e2e_enabled = True
            logger.debug(
                f"[MatrixParser] E2E Counter 识别: "
                f"{msg_info.name} → {sig_info.name}"
            )

        # 匹配 CRC 关键词
        elif any(kw in name_lower for kw in (
            "crc", "checksum", "chksum", "e2e_crc",
        )):
            sig_info.is_e2e_crc = True
            msg_info.e2e_crc_signal = sig_info.name
            msg_info.e2e_enabled = True
            logger.debug(
                f"[MatrixParser] E2E CRC 识别: "
                f"{msg_info.name} → {sig_info.name}"
            )

    @staticmethod
    def _extract_channel_name(sheet_name: str) -> str:
        """
        从 Sheet 名提取 Channel 名称。
        例如:
          'Tx_Public_CANFD1' → 'Public_CANFD1'
          'Rx_CCU_CANFD2'    → 'CCU_CANFD2'
          'Tx_TBOX_CANFD'    → 'TBOX_CANFD'
        """
        for prefix in ("Tx_", "Rx_"):
            if sheet_name.startswith(prefix):
                return sheet_name[len(prefix):]
        return sheet_name

    @staticmethod
    def _parse_msg_id(id_str: str) -> Optional[int]:
        """解析报文 ID 字符串 (支持 0x 前缀和纯数字)"""
        if not id_str:
            return None
        id_str = id_str.strip()
        try:
            if id_str.lower().startswith("0x"):
                return int(id_str, 16)
            else:
                return int(id_str)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _safe_str(value) -> str:
        """安全转换为字符串"""
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _safe_int(value, default: Optional[int] = None) -> Optional[int]:
        """安全转换为整数"""
        if value is None:
            return default
        try:
            return int(float(str(value)))
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _safe_float(
        value, default: Optional[float] = None
    ) -> Optional[float]:
        """安全转换为浮点数"""
        if value is None:
            return default
        try:
            return float(str(value))
        except (ValueError, TypeError):
            return default
